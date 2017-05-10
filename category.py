from openpyxl import load_workbook
import re
import os
import time

# 获取所有品类的内容已经熟悉
def getdata(path, targetPath):
    wb = load_workbook(path)
    sheet_names = wb.get_sheet_names()
    del sheet_names[1]  # 删除第二个sheet
    del sheet_names[0]  # 删除第一个sheet
    ListAll = []
    listMiss = []
    # 获取Excel文件数据组成列表
    for index in sheet_names:
        ws = wb.get_sheet_by_name(index)
        rowN = ws.max_row
        colN = ws.max_column

        ListTwo = []
        ListThree = []
        # 建立存储数据的字典
        data_dic = {}
        # 把数据存到字典中
        categoryOneName = {"category1Name": ws.cell(row=1, column=2).value}  # 获取一级品类名
        # 开始列循环 获取三级品类名字
        for ca3c in range(2, colN + 1):  # 从第二列开始循环
            temp_list = []
            categoryTwoValue = {"category2Value": None}
            for ca3r in range(3, rowN + 1):  # 从第三行开始循环
                w = ws.cell(row=ca3r, column=ca3c).value
                if w is not None:
                    categoryThreeName = {"category3Name": w}
                    # 查找目标路径
                    dataR = re.sub("[/,/]", "", w)
                    # 查找该品类属性文件对应路径
                    path = getpropertiespath(dataR, targetPath)
                    if path is None:
                        print("错误：查找不到名为【" + dataR + "】文件")
                        listMiss.append(dataR)
                        listprop = None
                    else:
                        # 读取该文件名的属性，并返回
                        listprop = getproperties(w, path)
                        if listprop == None:
                            listprop = None
                    categoryThreeValue = {"category3Value": listprop}
                    categoryThreeName.update(categoryThreeValue)  # 三级品类name，value合并
                    temp_list.append(categoryThreeName.copy())  # 加入list
                    categoryTwoValue = dict.fromkeys(categoryTwoValue, temp_list)  # 更新至对应二级品类value
            ListThree.append(categoryTwoValue.copy())  # 加入三级级品类list

        for ca2 in range(2, colN + 1):
            ca2Name = re.sub("[0-9]", "", ws.cell(row=2, column=ca2).value).strip()  # 去除二级品类数字和空格
            categoryTwoName = {"category2Name": ca2Name}
            ListTwo.append(categoryTwoName.copy())
        if len(ListTwo) == len(ListThree):
            for i in range(0, len(ListTwo)):
                ListTwo[i].update(ListThree[i])
        else:
            print('二级品类与三级品类不匹配:' + ws)
        categoryOneValue = {"category1Value": ListTwo.copy()}
        categoryOneName.update(categoryOneValue)
        ListAll.append(categoryOneName.copy())
    if len(listMiss) > 0:
        print("以下品类找不到属性文件" + str(listMiss))
    return ListAll


#  获取三级品类的文件全路径
def getpropertiespath(date, targetPath):
    if date is None:
        return print('date is none')
    date = date + '.xlsx'
    for r, ds, fs in os.walk(targetPath):
        for f in fs:
            fn = os.path.join(r, f)
            if f == date:
                propertiesPath = os.path.abspath(fn)
                return propertiesPath


# 获取三级品类的属性值
def getproperties(categoryName, path):
    wb = load_workbook(path)
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[0])
    rowN = ws.max_row
    colN = ws.max_column
    colNErr = None
    # 建立存储数据的字典
    data_dic = {}
    # 把数据存到字典中
    for rx in range(2, rowN + 1):
        temp_list = []
        pid = rx
        for rc in range(1, colN + 1):
            w = ws.cell(row=rx, column=rc).value
            temp_list.append(w)
        data_dic[pid] = temp_list
    propName = {"name"}
    listName = []
    propValue = {"value"}
    listValue = []
    indexErr = 0
    for name in data_dic.get(2):
        indexErr += 1
        if name == '属性':
            continue
        if name == None:
            print('提醒：属性名出现错误，请检查文件【{0}】第【{1}】列,可以删除该列 '\
                  .format(categoryName, indexErr))
            continue
        propName = dict.fromkeys(propName, name)
        listName.append(propName.copy())

    for rc in range(2, len(listName) + 2):
        temp_list = []
        for rx in range(3, rowN + 1):
            w = ws.cell(row=rx, column=rc).value
            temp_list.append(w)
        temp_list = list(set(temp_list))
        if temp_list.count(None) > 0:
            temp_list.remove(None)
        propValue = dict.fromkeys(propValue, temp_list)
        listValue.append(propValue.copy())
    if len(listName) == len(listValue):
        for i in range(0, len(listName)):
            listName[i].update(listValue[i])
    else:
        print('属性值与属性不匹配' + ws)
    return listName


# 生成SQL语句
def getsql(data,f):
    sql = 'INSERT INTO pms_product_category (' \
           'category_code, category_name, parent_category_code, ' \
           'category_comment, create_user, last_update_user, ' \
           'category_another_name,category_type_flag) VALUES (\''

    categoryOneCode = 0
    ISOTIMEFORMAT ='%Y-%m-%d %X'
    create_user = last_update_user = 'admin'
    category_comment = 'Null'
    create_date = last_update_date = time.strftime(ISOTIMEFORMAT, time.localtime(time.time()))
    parent_category_code = 'Null'
    category_another_name = 'Null'
    category_type_flag = 0
    if data == None:
        return print('获取不到生成SQL的数据')
    propid = 0
    for categoryOne in data:
        sqlOne1 = 'INSERT INTO pms_product_category (' \
               'category_code, category_name, create_user, last_update_user) VALUES (\''
        categoryOneCode += 1
        if categoryOneCode < 10:
            category_code1 = '0' + str(categoryOneCode)
        elif categoryOneCode <= 99:
            category_code1 = str(categoryOneCode)
        else:
            return print('第一品类数量大于99个，请重新计算')
        category_name1 = categoryOne.get('category1Name')
        sqlOne2 = '{0}\',\'{1}\',\'{2}\',\'{3}\');' \
            .format(category_code1, category_name1, create_user,
                    last_update_user)
        categoryTwoAll = categoryOne.get('category1Value')
        categoryTwoCode = 0
        print(sqlOne1 + sqlOne2, file=f)
        for categoryTwo in categoryTwoAll:
            sqlTwo1 = 'INSERT INTO pms_product_category (' \
               'category_code, category_name, parent_category_code, create_user, last_update_user) VALUES (\''
            categoryTwoCode += 1
            if categoryTwoCode < 10:
                category_code2 = category_code1 + '00' + str(categoryTwoCode)
            elif categoryTwoCode <= 99:
                category_code2 = category_code1 + '0' + str(categoryTwoCode)
            elif categoryTwoCode <= 999:
                category_code2 = category_code1 + str(categoryTwoCode)
            else:
                return print('第二品类数量大于999个，请重新计算')
            category_name2 = categoryTwo.get('category2Name')
            parent_category_code2 = category_code1
            sqlTwo2 = '{0}\',\'{1}\',\'{2}\',\'{3}\',\'{4}\');' \
                .format(category_code2, category_name2,parent_category_code2, create_user,
                        last_update_user)
            print(sqlTwo1 + sqlTwo2, file=f)
            categoryThreeAll = categoryTwo.get('category2Value')
            categoryThreeCode = 0
            for categoryThree in categoryThreeAll:
                sqlThree1 = 'INSERT INTO pms_product_category (' \
                   'category_code, category_name, parent_category_code, create_user, last_update_user) VALUES (\''
                categoryThreeCode += 1
                if categoryThreeCode < 10:
                    category_code3 = category_code2 + '00' + str(categoryThreeCode)
                elif categoryThreeCode <= 99:
                    category_code3 = category_code2 + '0' + str(categoryThreeCode)
                elif categoryThreeCode <= 999:
                    category_code3 = category_code2 + str(categoryThreeCode)
                else:
                    return print('第三品类数量大于999个，请重新计算')
                category_name3 = categoryThree.get('category3Name')
                parent_category_code3 = category_code2
                sqlThree2 = '{0}\',\'{1}\',\'{2}\',\'{3}\',\'{4}\');' \
                    .format(category_code3, category_name3, parent_category_code3, create_user,
                            last_update_user)
                print(sqlThree1 + sqlThree2, file=f)
                properties = categoryThree.get('category3Value')
                if properties is not None:
                    for prop in properties:
                        propid += 1
                        sqlprop1 = 'INSERT INTO pms_product_category_property (id, category_code, ' \
                                   'category_property_name, select_flag, create_user, last_update_user) VALUES (\''
                        id = propid
                        category_codeProp = category_code3
                        category_property_name = prop.get('name')
                        select_flag = '1'
                        sqlprop2 = '{0}\',\'{1}\',\'{2}\',\'{3}\',\'{4}\',\'{5}\');' \
                            .format(id, category_codeProp, category_property_name, select_flag,
                                    create_user,last_update_user)
                        print(sqlprop1 + sqlprop2, file=f)
                        propvalues = prop.get('value')
                        if propvalues is not None:
                            for propvalue in propvalues:
                                sqlpropvalue1 = 'INSERT INTO pms_product_category_property_value ' \
                                                '(category_property_id, category_property_value) VALUES (\''
                                category_property_id = id
                                if str(propvalue).find("'") > 0:
                                    propvalue = str(propvalue).replace("'", ".")
                                    print("提醒：表【pms_product_category_property_value】内【category_property_id】为【{0}】"
                                          "属性值【{1}】内含有错误标点【'】，以替换成【.】，"
                                          "如需更改请至数据库变更".format(category_property_id, propvalue))
                                category_property_value = propvalue
                                sqlpropvalue2 = '{0}\',\'{1}\');' \
                                    .format(category_property_id, category_property_value)
                                print(sqlpropvalue1 + sqlpropvalue2, file=f)

# 查询属性文件的上级路径
targetPath = r'F:\sp_doc\3.technology\1.PMS\3.Requirements\3.产品PRD文档\银犁食品产品品类+属性+属性值\银犁食品产品品类+属性+属性值-基础数据'
#  品类目录的文件路径
raadpath =r'F:\sp_doc\3.technology\1.PMS\3.Requirements\3.产品PRD文档\银犁食品产品品类+属性+属性值\银犁食品产品品类-分类最终版（订版及数据修改记录）--5个字符以内.xlsx'
#  输出SQL的文件路径
f = open("F:/sql.txt", "w+")
#  输出SQL到文件
getsql(getdata(raadpath, targetPath),f)
'''
getdata(raadpath, targetPath)
生成数据格式为：
[{'category1Name': '水果', 'category1Value': 
    [{'category2Name': '苹果类', 'category2Value': 
        [{'category3Name': '苹果', 'category3Value': 
            [{'name': '品种', 'value': ['爵士苹果', '蛇果/青蛇果', '花牛', '嘎啦果/加力果/姬娜果', '红玫瑰 ', '青苹果', '其它', '红富士', '爱妃', '黄金帅', '糖心苹果', '小苹果']}, 
             {'name': '品牌', 'value': ['都乐', '华圣']}, 
             {'name': '国产/进口', 'value': ['进口', '国产']}, 
             {'name': '产地', 'value': ['四川盐源', '美国', '陕西洛川', '新西兰', '山东蓬莱', '新疆阿克苏', '其它', '甘肃', '云南', '智利', '河北', '江苏', '山东烟台']}, 
             {'name': '等级', 'value': ['一等', '优等', 'Rose级', 'Beauty级', 'Queen级']}, 
             {'name': '果径', 'value': ['大果', '中果', '小果']}, 
             {'name': '规格', 'value': []}, 
             {'name': '包装方式', 'value': ['袋装', '礼盒装', '盒装']}, 
             {'name': '重量', 'value': []}]}]}]
'''
